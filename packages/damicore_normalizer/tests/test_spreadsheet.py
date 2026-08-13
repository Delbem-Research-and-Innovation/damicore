"""The spreadsheet dataset source and the cell-text rule it depends on.

The rule is asserted per cell type rather than through a whole-file comparison, because it
is what object bytes -- and therefore every distance -- are a function of. A rule that
stringified a value plausibly but differently would produce artifacts that pass every
structural check while measuring something the user did not supply.
"""

from __future__ import annotations

import datetime as dt
import json
import tracemalloc
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from damicore_normalizer import (
    DelimitedSource,
    NormalizationConfig,
    NormalizerError,
    SpreadsheetSource,
    materialize_objects,
)
from damicore_normalizer.spreadsheet_reader import CellValue, cell_text

pytestmark = pytest.mark.unit

SHEET = NormalizationConfig(source=SpreadsheetSource())


def _workbook(path: Path, rows: list[list[CellValue]], *, title: str = "Sheet") -> Path:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = title
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    return path


# Every branch of cell_text_rule v1, as (cell value, the text it must produce). The integral
# float rows are the ones that make a spreadsheet and a delimited file agree: a reader that
# rendered 42.0 would produce different object bytes for the same logical table.
@pytest.mark.parametrize(
    ("value", "expected"),
    [
        pytest.param("plain", "plain", id="text"),
        pytest.param("", "", id="empty-string"),
        pytest.param(None, "", id="blank"),
        pytest.param(True, "TRUE", id="true"),
        pytest.param(False, "FALSE", id="false"),
        pytest.param(42, "42", id="integer"),
        pytest.param(42.0, "42", id="integral-float"),
        pytest.param(-0.0, "0", id="negative-zero"),
        pytest.param(1e16, "10000000000000000", id="large-integral-float"),
        pytest.param(3.5, "3.5", id="fractional-float"),
        pytest.param(0.1 + 0.2, "0.30000000000000004", id="float-keeps-round-trip-repr"),
        pytest.param(dt.date(2026, 1, 2), "2026-01-02T00:00:00", id="date"),
        pytest.param(dt.datetime(2026, 1, 2, 13, 45, 30), "2026-01-02T13:45:30", id="datetime"),
        pytest.param(dt.time(13, 45, 30), "13:45:30", id="time"),
        pytest.param("#REF!", "#REF!", id="error-value-is-text"),
        pytest.param("=1+1", "=1+1", id="formula-text"),
    ],
)
def test_cell_text_rule_v1_renders_each_cell_type(value: CellValue, expected: str) -> None:
    assert cell_text(value) == expected


def test_cell_text_rule_refuses_a_non_finite_number() -> None:
    """A cell holding infinity has no decimal rendering, and `allow_nan=False` would reject it
    at the manifest boundary anyway -- later, and with a worse message."""
    with pytest.raises(NormalizerError, match="non-finite number") as raised:
        cell_text(float("inf"))
    assert raised.value.code == "dataset_format_error"


def test_cell_text_rule_refuses_a_value_it_does_not_span(tmp_path: Path) -> None:
    """Totality is the point: a type outside the rule must fail loudly rather than become a
    plausible string nobody chose."""
    with pytest.raises(NormalizerError, match="unsupported value of type timedelta") as raised:
        cell_text(dt.timedelta(hours=1))
    assert raised.value.code == "dataset_format_error"


def test_a_workbook_and_a_delimited_file_of_one_table_produce_the_same_bytes(
    tmp_path: Path,
) -> None:
    """The parity that makes the format an implementation detail of the input rather than a
    property of the result."""
    table: list[list[CellValue]] = [
        ["alpha", "beta", "gamma"],
        ["text", 1, 2.5],
        ["ünicode ✓", 10, 0.5],
    ]
    _workbook(tmp_path / "book.xlsx", table)
    (tmp_path / "table.csv").write_text(
        "alpha,beta,gamma\ntext,1,2.5\nünicode ✓,10,0.5\n", encoding="utf-8"
    )

    from_sheet = materialize_objects(tmp_path / "book.xlsx", tmp_path / "sheet", config=SHEET)
    from_text = materialize_objects(
        tmp_path / "table.csv",
        tmp_path / "text",
        config=NormalizationConfig(source=DelimitedSource()),
    )

    assert [item.sha256 for item in from_sheet.objects] == [
        item.sha256 for item in from_text.objects
    ]
    assert [item.label for item in from_sheet.objects] == ["alpha", "beta", "gamma"]


def test_formatting_outside_the_data_does_not_invent_objects(tmp_path: Path) -> None:
    """openpyxl reports a sheet's dimensions from anything it stores, so one formatting-only
    cell far below the data inflates them. Untrimmed, those phantom rows become empty objects
    that are mutually identical and cluster together, which is a wrong answer that every
    structural check accepts."""
    path = tmp_path / "ragged.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["a", "b", "c"])
    sheet.append(["1", "2", "3"])
    sheet.append(["4", "5", "6"])
    sheet["H40"].fill = PatternFill("solid", fgColor="FFFF00")
    workbook.save(path)

    columns = materialize_objects(path, tmp_path / "columns", config=SHEET)
    assert [item.label for item in columns.objects] == ["a", "b", "c"]

    rows = materialize_objects(
        path,
        tmp_path / "rows",
        config=NormalizationConfig(source=SpreadsheetSource(split="rows")),
    )
    assert rows.object_count == 2


def test_a_multi_sheet_workbook_requires_an_explicit_choice(tmp_path: Path) -> None:
    """Defaulting to the first sheet would silently decide which data was analyzed."""
    path = tmp_path / "two.xlsx"
    workbook = Workbook()
    first = workbook.active
    assert first is not None
    first.title = "First"
    first.append(["a", "b"])
    first.append(["1", "2"])
    second = workbook.create_sheet("Second")
    second.append(["c", "d"])
    second.append(["3", "4"])
    workbook.save(path)

    with pytest.raises(NormalizerError, match="one must be named") as raised:
        materialize_objects(path, tmp_path / "out", config=SHEET)
    assert raised.value.code == "dataset_format_error"

    chosen = materialize_objects(
        path,
        tmp_path / "second",
        config=NormalizationConfig(source=SpreadsheetSource(sheet="Second")),
    )
    assert [item.label for item in chosen.objects] == ["c", "d"]
    manifest = json.loads(chosen.manifest_path.read_text(encoding="utf-8"))
    assert manifest["input"]["sheet"] == "Second"
    assert manifest["input"]["cell_text_rule"] == "v1"
    assert manifest["input"]["kind"] == "xlsx"


def test_an_unknown_sheet_name_is_refused_with_the_available_ones(tmp_path: Path) -> None:
    path = _workbook(tmp_path / "one.xlsx", [["a", "b"], ["1", "2"]], title="Only")
    with pytest.raises(NormalizerError, match="available: Only") as raised:
        materialize_objects(
            path, tmp_path / "out", config=NormalizationConfig(source=SpreadsheetSource(sheet="No"))
        )
    assert raised.value.code == "dataset_format_error"


def test_a_formula_is_stored_as_text_and_never_evaluated(tmp_path: Path) -> None:
    """AGENTS.md: input is data, never code. Reading the cached value instead would return
    nothing for a workbook the spreadsheet application never recalculated, turning a formula
    into a blank without saying so."""
    path = _workbook(
        tmp_path / "formula.xlsx",
        [["a", "b"], ["1", "=1+1"], ["2", "=SUM(A2:A3)"]],
    )
    result = materialize_objects(path, tmp_path / "out", config=SHEET)
    payload = (tmp_path / "out" / result.objects[1].relative_path).read_bytes()
    assert payload == b'"=1+1"\n"=SUM(A2:A3)"\n'


def test_a_merged_cell_is_read_as_presented(tmp_path: Path) -> None:
    """The value sits in the leading cell and the rest are blank. Rejecting a merge is
    hostile and filling it invents data the workbook does not contain."""
    path = tmp_path / "merged.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["a", "b", "c"])
    sheet.append(["merged", None, "3"])
    sheet.append(["4", "5", "6"])
    sheet.merge_cells("A2:B2")
    workbook.save(path)

    result = materialize_objects(path, tmp_path / "out", config=SHEET)
    assert (tmp_path / "out" / result.objects[1].relative_path).read_bytes() == b'""\n"5"\n'


@pytest.mark.parametrize(
    ("name", "payload", "discriminator"),
    [
        pytest.param("legacy.xls", b"\xd0\xcf\x11\xe0", "Convert it to .xlsx", id="legacy-xls"),
        pytest.param("broken.xlsx", b"not a zip container", "Could not open", id="not-a-workbook"),
        pytest.param("empty.xlsx", b"", "Could not open", id="empty-file"),
    ],
)
def test_a_workbook_that_cannot_be_read_is_refused(
    tmp_path: Path, name: str, payload: bytes, discriminator: str
) -> None:
    path = tmp_path / name
    path.write_bytes(payload)
    with pytest.raises(NormalizerError, match=discriminator) as raised:
        materialize_objects(path, tmp_path / "out", config=SHEET)
    assert raised.value.code == "dataset_format_error"


def test_a_worksheet_with_no_data_is_refused(tmp_path: Path) -> None:
    path = tmp_path / "blank.xlsx"
    Workbook().save(path)
    with pytest.raises(NormalizerError, match="contains no data") as raised:
        materialize_objects(path, tmp_path / "out", config=SHEET)
    assert raised.value.code == "dataset_format_error"


def _timed_peak(tmp_path: Path, rows: int, label: str) -> int:
    """Build a sheet of `rows` rows drawn from a small vocabulary and measure the peak."""
    path = tmp_path / f"{label}.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    vocabulary = [f"token_{index:02d}" for index in range(16)]
    sheet.append([f"column_{index}" for index in range(6)])
    for row in range(rows):
        sheet.append([vocabulary[(row + index) % len(vocabulary)] for index in range(6)])
    workbook.save(path)

    tracemalloc.start()
    materialize_objects(path, tmp_path / f"out-{label}", config=SHEET)
    peak = tracemalloc.get_traced_memory()[1]
    tracemalloc.stop()
    return peak


def test_reading_a_worksheet_does_not_scale_memory_with_its_row_count(tmp_path: Path) -> None:
    """The bounded-memory invariant for this source.

    `pandas.read_excel` has no chunksize and would hold the whole sheet, which is why the
    reader goes through openpyxl's row iterator instead. The property asserted is the one
    that actually holds: peak memory is governed by the workbook's shared-string table and
    not by how many rows reference it, so a tenfold longer sheet must not cost tenfold. The
    fixture draws from a fixed vocabulary precisely so the string table stays constant while
    the row count does not.
    """
    small = _timed_peak(tmp_path, 500, "small")
    large = _timed_peak(tmp_path, 5_000, "large")

    # Ten times the rows. Streaming keeps the growth far below that; materializing the sheet
    # would track it. The bound is loose on purpose -- the claim is about the exponent, not
    # about a particular allocator's footprint.
    assert large < small * 3, (small, large)
