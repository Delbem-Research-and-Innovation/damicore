from __future__ import annotations

import datetime as dt
import math
import zipfile
from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException

from damicore_normalizer.config import SpreadsheetSource
from damicore_normalizer.errors import NormalizerError
from damicore_normalizer.table_split import TableScan, split_table, validate_table_shape

CELL_TEXT_RULE = "v1"

# The closed set of Python types a worksheet read with ``values_only=True`` can yield. It is
# what makes ``cell_text`` total: the coercion has a finite domain to span, so a type outside
# it is a static error here rather than an unhandled value at runtime.
CellValue = str | int | float | bool | dt.datetime | dt.date | dt.time | dt.timedelta | None

# openpyxl raises these when it cannot make sense of a workbook: a file that is not a zip
# container (BadZipFile, which derives straight from Exception rather than OSError), an
# encrypted one, or a member the reader cannot parse. They are caught as a set so a malformed
# workbook is one typed failure rather than whichever library exception happened to surface.
_WORKBOOK_FAILURES = (
    OSError,
    ValueError,
    KeyError,
    TypeError,
    zipfile.BadZipFile,
    InvalidFileException,
)


@dataclass(frozen=True)
class _UsedRange:
    """The smallest rectangle containing every non-blank cell, as 1-based inclusive bounds."""

    min_row: int
    max_row: int
    min_column: int
    max_column: int


def cell_text(value: CellValue) -> str:
    """Render one spreadsheet cell as text under ``cell_text_rule`` v1.

    The rule, not the parsing library, is what object bytes depend on, so it lives here and
    is named in the manifest. Two properties matter and both are load-bearing. It is total:
    every type openpyxl can yield is handled, and an unhandled one raises rather than
    stringifying into something plausible. And it is parity-preserving: an integral value
    renders without a fractional part, so the same logical table read from a spreadsheet and
    from delimited text produces identical object bytes.

    Raises
    ------
    NormalizerError
        The cell holds a value this rule does not span (``dataset_format_error``).
    """
    if value is None:
        # A blank cell and an empty string are indistinguishable through openpyxl, and a
        # blank delimited field already produces "", so one value spans all three.
        return ""
    # bool is a subclass of int, so it has to be decided before the numeric branches.
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, str):
        return value
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            raise NormalizerError(
                "Spreadsheet cell holds a non-finite number",
                code="dataset_format_error",
            )
        return str(int(value)) if value.is_integer() else repr(value)
    # datetime is a subclass of date, so the narrower type is decided first. A date cell
    # reaches here as a midnight datetime -- openpyxl has already lost the distinction -- so
    # both render with a time component rather than depending on the cell's number format,
    # which would make object bytes change when a cell is reformatted but not edited.
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%dT%H:%M:%S")
    if isinstance(value, dt.date):
        return f"{value.isoformat()}T00:00:00"
    if isinstance(value, dt.time):
        return value.strftime("%H:%M:%S")
    raise NormalizerError(
        f"Spreadsheet cell holds an unsupported value of type {type(value).__name__}",
        code="dataset_format_error",
    )


def _open(path: Path) -> Workbook:
    if path.suffix.lower() == ".xls":
        # Named before openpyxl reports it as an unreadable container, because the useful
        # answer is a one-step conversion rather than a parse failure. Legacy BIFF is out of
        # scope; see docs/decisions/0009-spreadsheet-engine.md.
        raise NormalizerError(
            f"Legacy .xls workbooks are not supported: {path.name}. Convert it to .xlsx and "
            "read that instead.",
            code="dataset_format_error",
        )
    try:
        # data_only=False on purpose: the cached alternative returns nothing for a workbook
        # the spreadsheet application never recalculated, which would turn a formula into a
        # blank without saying so. Formula text is deterministic and is never evaluated.
        return openpyxl.load_workbook(path, read_only=True, data_only=False)
    except _WORKBOOK_FAILURES as exc:
        raise NormalizerError(
            f"Could not open workbook: {path.name}",
            code="dataset_format_error",
        ) from exc


def resolve_sheet(path: Path, source: SpreadsheetSource) -> str:
    """Name the worksheet to read, refusing to guess when a workbook holds several."""
    workbook = _open(path)
    try:
        names = list(workbook.sheetnames)
    finally:
        workbook.close()
    if not names:
        raise NormalizerError("Workbook contains no worksheet", code="dataset_format_error")
    if source.sheet is None:
        if len(names) > 1:
            raise NormalizerError(
                f"Workbook contains {len(names)} worksheets, so one must be named; "
                f"available: {', '.join(names)}",
                code="dataset_format_error",
            )
        return names[0]
    if source.sheet not in names:
        raise NormalizerError(
            f"Workbook has no worksheet named {source.sheet!r}; available: {', '.join(names)}",
            code="dataset_format_error",
        )
    return source.sheet


def _used_range(path: Path, sheet: str) -> _UsedRange:
    """Find the real data rectangle, ignoring cells that carry only formatting.

    openpyxl reports a sheet's dimensions from anything it stores, so a single filled cell
    far below the data inflates them; a 2x3 sheet with formatting at H40 is reported as
    40x8. Left alone those phantom cells become objects -- empty, mutually identical, and a
    cluster that means nothing. One streaming pass establishes the bounds, and the reading
    pass below slices to them.
    """
    minimum_row = 0
    maximum_row = 0
    minimum_column = 0
    maximum_column = 0
    workbook = _open(path)
    try:
        worksheet = workbook[sheet]
        for number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            filled = [index for index, value in enumerate(row, start=1) if value is not None]
            if not filled:
                continue
            first, last = filled[0], filled[-1]
            minimum_row = minimum_row or number
            maximum_row = number
            minimum_column = min(minimum_column or first, first)
            maximum_column = max(maximum_column, last)
    except _WORKBOOK_FAILURES as exc:
        raise NormalizerError("Workbook parsing failed", code="dataset_format_error") from exc
    finally:
        workbook.close()
    if minimum_row == 0:
        raise NormalizerError("Worksheet contains no data", code="dataset_format_error")
    return _UsedRange(minimum_row, maximum_row, minimum_column, maximum_column)


def _iter_used_rows(path: Path, sheet: str, bounds: _UsedRange) -> Iterator[tuple[str, ...]]:
    workbook = _open(path)
    try:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(
            min_row=bounds.min_row,
            max_row=bounds.max_row,
            min_col=bounds.min_column,
            max_col=bounds.max_column,
            values_only=True,
        ):
            yield tuple(cell_text(value) for value in row)
    finally:
        workbook.close()


def scan_spreadsheet(
    path: Path,
    source: SpreadsheetSource,
    *,
    chunk_rows: int,
    max_open_files: int,
    objects_dir: Path | None = None,
) -> tuple[TableScan, str]:
    """Scan one worksheet into canonical objects; optionally persist them.

    Returns the scan together with the resolved worksheet name, which the manifest records
    so a completed run never leaves which sheet was analyzed to be inferred.
    """
    sheet = resolve_sheet(path, source)
    bounds = _used_range(path, sheet)
    rows = _iter_used_rows(path, sheet, bounds)
    try:
        header = next(rows)
    except StopIteration as exc:  # pragma: no cover - _used_range already rejects this
        raise NormalizerError("Worksheet contains no data", code="dataset_format_error") from exc
    validate_table_shape(header, source.split)

    if objects_dir is not None:
        objects_dir.mkdir(parents=True, exist_ok=False)
    try:
        scan = split_table(
            header,
            rows,
            split=source.split,
            chunk_rows=chunk_rows,
            max_open_files=max_open_files,
            objects_dir=objects_dir,
        )
    except NormalizerError:
        raise
    except _WORKBOOK_FAILURES as exc:
        raise NormalizerError("Workbook parsing failed", code="dataset_format_error") from exc
    return scan, sheet
